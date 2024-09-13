import tkinter as tk
from tkinter import Tk, messagebox
from tkinter.ttk import Treeview
from openpyxl import Workbook, load_workbook
from tkinter.constants import NSEW
import os, re
from tkinter import Label, Entry, Button

def tabla_contactos():
    
    # Crear la ventana principal
    ventana = Tk()
    ventana.title("Lista de Contactos")
    ventana.geometry("600x400")

    # Crear el Treeview
    tree = Treeview(ventana, columns=("Nombre", "Apellido", "Telefono", "Email", "Apodo", "Direccion"), show='headings')

    # Configurar las columnas
    tree.column("Nombre", anchor='center', width=100)
    tree.column("Apellido", anchor='center', width=100)
    tree.column("Telefono", anchor='center', width=100)
    tree.column("Email", anchor='center', width=150)
    tree.column("Apodo", anchor="center", width=80)
    tree.column("Direccion", anchor="center", width=150)

    # Definir los encabezados
    tree.heading("Nombre", text="Nombre")
    tree.heading("Apellido", text="Apellido")
    tree.heading("Telefono", text="Telefono")
    tree.heading("Email", text="Email")
    tree.heading("Apodo", text="Apodo")
    tree.heading("Direccion", text="Direccion")

    # Colocar el Treeview en la ventana
    tree.grid(row=0, column=0, columnspan=3, sticky=NSEW)

    # Crear etiquetas y campos de entrada
    Label(ventana, text="Nombre: ").grid(row=1, column=0)
    Label(ventana, text="Apellido: ").grid(row=2, column=0)
    Label(ventana, text="Telefono: ").grid(row=3, column=0)
    Label(ventana, text="Email: ").grid(row=4, column=0)
    Label(ventana, text="Apodo: ").grid(row=5, column=0)
    Label(ventana, text="Direccion: ").grid(row=6, column=0)

    nombre_entry = Entry(ventana)
    apellido_entry = Entry(ventana)
    telefono_entry = Entry(ventana)
    email_entry = Entry(ventana)
    apodo_entry = Entry(ventana)
    direccion_entry = Entry(ventana)

    nombre_entry.grid(row=1, column=1)
    apellido_entry.grid(row=2, column=1)
    telefono_entry.grid(row=3, column=1)
    email_entry.grid(row=4, column=1)
    apodo_entry.grid(row=5, column=1)
    direccion_entry.grid(row=6, column=1)

    # Función para exportar los datos a Excel
    def funcion_exportar():
        nombre_archivo = r'datos.xlsx'

        # Comprobar si el archivo ya existe
        if os.path.exists(nombre_archivo):
            wb = load_workbook(nombre_archivo)
            ws = wb.active
        else:
            # Crear el libro Excel si no existe
            wb = Workbook()
            ws = wb.active
            ws.append(["Apellido", "Nombre", "Apodo", "Telefono", "Email", "Dirección"])

        # Agregar los datos al archivo Excel
        ws.append([apellido_entry.get(), nombre_entry.get(), apodo_entry.get(), telefono_entry.get(), email_entry.get(), direccion_entry.get()])
        wb.save(nombre_archivo)
        messagebox.showinfo(title="Información", message="Datos guardados con éxito")

    # Función para guardar los datos (Create)
    def guardar_datos():
        nombre = nombre_entry.get()
        apellido = apellido_entry.get()
        apodo = apodo_entry.get()
        telefono = telefono_entry.get()
        email = email_entry.get()
        direccion = direccion_entry.get()

        # Validar campos vacíos
        if not apellido or not nombre or not apodo or not telefono or not email or not direccion:
            messagebox.showwarning(title="Advertencia", message="Todos los campos son obligatorios")
            return

        # Validar que el teléfono sea un número
        try:
            telefono = int(telefono)
        except ValueError:
            messagebox.showwarning(title="Advertencia", message="El teléfono debe ser un número")
            return

        # Validar formato de email
        if not re.match(r"[^@]+@[^@]+\.[^@]+", email):
            messagebox.showwarning(title="Advertencia", message="Correo electrónico no válido")
            return

        # Agregar los datos al Treeview
        tree.insert("", "end", values=(nombre, apellido, telefono, email, apodo, direccion))
        
        # Llamar a la función para exportar los datos
        funcion_exportar()

        # Limpiar los campos de entrada
        nombre_entry.delete(0, tk.END)
        apellido_entry.delete(0, tk.END)
        apodo_entry.delete(0, tk.END)
        telefono_entry.delete(0, tk.END)
        email_entry.delete(0, tk.END)
        direccion_entry.delete(0, tk.END)

        # Llamar a la función para exportar los datos
        funcion_exportar()

    # Función para cargar datos seleccionados al formulario (Read)
    def cargar_datos():
        selected_item = tree.selection()
        if not selected_item:
            messagebox.showwarning(title="Advertencia", message="Por favor selecciona un contacto.")
            return

        item = tree.item(selected_item)
        valores = item["values"]

        # Cargar los valores en los campos de entrada
        nombre_entry.delete(0, tk.END)
        apellido_entry.delete(0, tk.END)
        telefono_entry.delete(0, tk.END)
        email_entry.delete(0, tk.END)
        apodo_entry.delete(0, tk.END)
        direccion_entry.delete(0, tk.END)

        nombre_entry.insert(0, valores[0])
        apellido_entry.insert(0, valores[1])
        telefono_entry.insert(0, valores[2])
        email_entry.insert(0, valores[3])
        apodo_entry.insert(0, valores[4])
        direccion_entry.insert(0, valores[5])

    # Función para actualizar un contacto seleccionado (Update)
    def actualizar_datos():
        selected_item = tree.selection()
        if not selected_item:
            messagebox.showwarning(title="Advertencia", message="Por favor selecciona un contacto.")
            return

        tree.item(selected_item, values=(nombre_entry.get(), apellido_entry.get(), telefono_entry.get(), email_entry.get(), apodo_entry.get(), direccion_entry.get()))
        
        messagebox.showinfo(title="Información", message="Contacto actualizado con éxito.")

    # Función para eliminar un contacto seleccionado (Delete)
    def eliminar_datos():
        selected_item = tree.selection()
        if not selected_item:
            messagebox.showwarning(title="Advertencia", message="Por favor selecciona un contacto.")
            return

        tree.delete(selected_item)
        messagebox.showinfo(title="Información", message="Contacto eliminado con éxito.")

    # Botones para CRUD
    boton_guardar = tk.Button(ventana, text="Agregar Contacto", command=guardar_datos, bg='#6D8299', fg='white')
    boton_guardar.grid(row=7, column=0, padx=10, pady=10)

    boton_cargar = tk.Button(ventana, text="Cargar a Excel", command=cargar_datos, bg='#6D8299', fg='white')
    boton_cargar.grid(row=7, column=1, padx=10, pady=10)

    boton_actualizar = tk.Button(ventana, text="Actualizar Contacto", command=actualizar_datos, bg='#6D8299', fg='white')
    boton_actualizar.grid(row=8, column=0, padx=10, pady=10)

    boton_eliminar = tk.Button(ventana, text="Eliminar Contacto", command=eliminar_datos, bg='#6D8299', fg='white')
    boton_eliminar.grid(row=8, column=1, padx=10, pady=10)

    # Configurar el layout de la ventana para que el Treeview se redimensione correctamente
    ventana.grid_rowconfigure(0, weight=1)
    ventana.grid_columnconfigure(0, weight=1)
    ventana.grid_columnconfigure(2, weight=1)

    # Iniciar el bucle principal de la aplicación
    ventana.mainloop()