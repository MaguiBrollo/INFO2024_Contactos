#=======IMPORTACIONES=============================================

import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from PIL import ImageTk, Image
import re
import os
import contactos 
import pandas as pd

#======FUNCIONES==============================================
def funcion_login():
   messagebox.showwarning("Login", "Aquí se hace Inicio de Sesión")

def funcion_logout():
   messagebox.showwarning("Logout", "Aqui se hace CERRAR sesión")

def mostrar_contactos():
   messagebox.showwarning("Contactos", "Aquí se muestra todos los contactos, y los botones Crear/Eliminar/Editar") 

def funcion_exportar():
   nombre_archivo = 'datos.xlsx'

   # Comprobar si el archivo ya existe
   if os.path.exists(nombre_archivo):
      wb = load_workbook(nombre_archivo)
      ws = wb.active
   else:
      # Crear el libro excel
      wb = Workbook()
      ws = wb.active
      ws.append(["Apellido", "Nombre", "Apodo", "Telefono", "Email", "Dirección"])
   def guardar_datos():
      apellido = entry_apellido.get()
      nombre = entry_nombre.get()
      apodo = entry_apodo.get()
      telefono = entry_telefono.get()
      email = entry_email.get()
      direccion = entry_direccion.get()

      if not apellido or not nombre or not apodo or not telefono or not email or not direccion:
         messagebox.showwarning(title="Advertencia", message="Todos los campos son obligatorios")
         return
      try:
         telefono = int(telefono)
      except ValueError:
         messagebox.showwarning(title="Advertencia", message="telefono debe ser numero")
         return
      
      # Validar formato de email
      if not re.match(r"[^@]+@[^@]+\.[^@]+", email):
         messagebox.showwarning(title="Advertencia", message="Correo electrónico no válido")
         return

      ws.append([apellido, nombre, apodo, telefono, email, direccion])
      wb.save(nombre_archivo)
      messagebox.showinfo(title="Información", message="Datos guardados con éxito")

      entry_apellido.delete(0, tk.END)
      entry_nombre.delete(0, tk.END)
      entry_apodo.delete(0, tk.END)
      entry_telefono.delete(0, tk.END)
      entry_email.delete(0, tk.END)
      entry_direccion.delete(0, tk.END)



   root = tk.Tk()
   root.title("Lista de Entrada de Datos")
   root.configure(bg='#171')
   label_style = {"bg": '#171', "fg": "white"}
   entry_style = {"bg": '#D3D3D3', "fg": "black"}

   label_apellido = tk.Label(root, text="Apellido", **label_style)
   label_apellido.grid(row=0, column=0, padx=10, pady=5)
   entry_apellido = tk.Entry(root, **entry_style)
   entry_apellido.grid(row=0, column=1, padx=10, pady=5)

   label_nombre = tk.Label(root, text="Nombre", **label_style)
   label_nombre.grid(row=1, column=0, padx=10, pady=5)
   entry_nombre = tk.Entry(root, **entry_style)
   entry_nombre.grid(row=1, column=1, padx=10, pady=5)

   label_apodo = tk.Label(root, text="Apodo", **label_style)
   label_apodo.grid(row=2, column=0, padx=10, pady=5)
   entry_apodo = tk.Entry(root, **entry_style)
   entry_apodo.grid(row=2, column=1, padx=10, pady=5)

   label_telefono = tk.Label(root, text="Telefono", **label_style)
   label_telefono.grid(row=3, column=0, padx=10, pady=5)
   entry_telefono = tk.Entry(root, **entry_style)
   entry_telefono.grid(row=3, column=1, padx=10, pady=5)

   label_email = tk.Label(root, text="Email", **label_style)
   label_email.grid(row=4, column=0, padx=10, pady=5)
   entry_email = tk.Entry(root, **entry_style)
   entry_email.grid(row=4, column=1, padx=10, pady=5)

   label_direccion = tk.Label(root, text="Direccion", **label_style)
   label_direccion.grid(row=5, column=0, padx=10, pady=5)
   entry_direccion = tk.Entry(root, **entry_style)
   entry_direccion.grid(row=5, column=1, padx=10, pady=5)

   boton_guardar = tk.Button(root, text="Guardar", command=guardar_datos, bg='#6D8299', fg='white')
   boton_guardar.grid(row=6, column=0, columnspan=2, padx=10, pady=10)
   root.mainloop()   
      #messagebox.showwarning("Exportar", "Aquí la opción de Exportar los contactos a EXCEL")



#=====PROGRAMA PRINCIPAL===============================================
ventana = tk.Tk()
lista_contactos = tk.Listbox(ventana)
login = False

dir_carpetas = os.path.dirname(__file__)
img_carpeta = os.path.join(dir_carpetas,"imagenes")
ventana.iconbitmap(os.path.join(img_carpeta,"contactos.ico"))

ventana.title('Agenda Personal')
ventana.geometry('800x600')
ventana.configure(bg="#4B6587")

img_logo_ppal = ImageTk.PhotoImage(Image.open(os.path.join(img_carpeta,"contactos.png")).resize((100,100)))
img_mostrar = tk.Label(image=img_logo_ppal)
img_mostrar.place(relx=0.8,rely=0.8)

barra_menu = tk.Menu(ventana)
ventana.config(menu=barra_menu)
menu_principal = tk.Menu(barra_menu)
barra_menu.add_cascade(label ='Menú', menu=menu_principal)
submenu = tk.Menu(menu_principal)

if(login):
   menu_principal.add_cascade(label = 'Login', command=funcion_login)
else:
   menu_principal.add_cascade(label = 'Contactos', menu=submenu)
   menu_principal.add_cascade(label = 'Logout', command=funcion_logout)

submenu.add_cascade(label = 'Actualizar Contactos',command=mostrar_contactos)
submenu.add_cascade(label = 'Exportar a Excel', command=funcion_exportar) 

ventana.mainloop()